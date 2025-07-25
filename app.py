from flask import Flask, render_template, request, send_file, redirect, url_for, flash, jsonify
import os
import uuid
import random
import time
from openpyxl import Workbook
from werkzeug.utils import secure_filename
import zipfile

app = Flask(__name__)
app.secret_key = 'your-secret-key'

UPLOAD_FOLDER = 'uploads'
OUTPUT_FOLDER = 'outputs'

os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(OUTPUT_FOLDER, exist_ok=True)

@app.route('/', methods=['GET', 'POST'])
def index():
    return render_template('index.html')

@app.route('/help')
def help_page():
    return render_template('help.html')

@app.route('/count_ids', methods=['POST'])
def count_ids():
    try:
        file = request.files.get('id_file')
        if not file:
            return jsonify({"error": "No file uploaded"}), 400

        ext = os.path.splitext(file.filename)[1].lower()
        count = 0

        if ext == '.txt':
            for i, line in enumerate(file.stream):
                if i > 10000:
                    break
                if line.strip():
                    count += 1

        elif ext in ['.xlsx', '.xls']:
            import openpyxl
            from io import BytesIO
            wb = openpyxl.load_workbook(filename=BytesIO(file.read()), read_only=True)
            sheet = wb.active
            for i, row in enumerate(sheet.iter_rows(min_row=1, max_col=1), 1):
                if i > 10000:
                    break
                val = row[0].value
                if val:
                    count += 1
        else:
            return jsonify({"error": "Unsupported file format"}), 400

        return jsonify({"count": count})

    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/generate', methods=['POST'])
def generate():
    try:
        length = request.form.get('length')
        count_raw = request.form.get('count')
        if not count_raw or not count_raw.isdigit():
            flash("Please enter a valid number of links.")
            return redirect(url_for('index'))
        count = int(count_raw)

        base_url = request.form.get('base_url').strip()
        prefix = request.form.get('prefix', '')
        include_ptest = request.form.get('ptest') == 'on'
        generate_test = request.form.get('generate_test') == 'on'
        test_count = max(1, int(request.form.get('test_count', 20)))

        use_uuid = not length.isdigit() if length else True
        id_length = int(length) if not use_uuid else 0

        ids = []
        file = request.files.get('id_file')
        if file and file.filename:
            filename = secure_filename(file.filename)
            filepath = os.path.join(UPLOAD_FOLDER, filename)
            file.save(filepath)
            ext = os.path.splitext(filename)[1].lower()
            if ext == '.txt':
                with open(filepath, 'r') as f:
                    ids = [line.strip() for line in f if line.strip()]
            elif ext in ['.xlsx', '.xls']:
                import openpyxl
                wb = openpyxl.load_workbook(filepath, read_only=True)
                sheet = wb.active
                for row in sheet.iter_rows(min_row=1, max_col=1):
                    val = row[0].value
                    if val:
                        ids.append(str(val).strip())
            if count > len(ids):
                flash("Not enough IDs in file.")
                return redirect(url_for('index'))
            ids = ids[:count]
        else:
            seen = set()
            while len(ids) < count:
                new_id = str(uuid.uuid4()).replace('-', '') if use_uuid else ''.join(random.choices('0123456789ABCDEFGHIJKLMNOPQRSTUVWXYZ', k=id_length))
                if new_id not in seen:
                    seen.add(new_id)
                    ids.append(new_id)

        timestamp = time.strftime("%Y%m%d_%H%M%S")
        filename_prefix = f"{count}_Unique_{timestamp}"
        id_file_path = os.path.join(OUTPUT_FOLDER, f"{filename_prefix}.dat")
        xlsx_file_path = os.path.join(OUTPUT_FOLDER, f"{filename_prefix}.xlsx")

        # Efficient write for large volumes
        from openpyxl import Workbook
        wb = Workbook(write_only=True)
        ws = wb.create_sheet(title="Links")
        ws.append(["ID", "Link"])

        with open(id_file_path, 'w') as f1:
            for uid in ids:
                full_id = prefix + uid
                url = f"{base_url}{full_id}"
                if include_ptest:
                    url += "&ptest=0"
                f1.write(full_id + '\n')
                ws.append([full_id, url])

        wb.save(xlsx_file_path)

        test_file_path = None
        if generate_test:
            test_file_path = os.path.join(OUTPUT_FOLDER, f"TestLinks_{timestamp}.xlsx")
            test_wb = Workbook(write_only=True)
            test_ws = test_wb.create_sheet(title="Test Links")
            test_ws.append(["ID", "Link"])
            for i in range(1, test_count + 1):
                test_id = f"TEST_{i}"
                test_url = f"{base_url}{test_id}"
                if include_ptest:
                    test_url += "&ptest=0"
                test_ws.append([test_id, test_url])
            test_wb.save(test_file_path)

        # Zip outputs
        zip_filename = f"{filename_prefix}_all_files.zip"
        zip_path = os.path.join(OUTPUT_FOLDER, zip_filename)

        with zipfile.ZipFile(zip_path, 'w') as zipf:
            zipf.write(id_file_path, arcname=os.path.basename(id_file_path))
            zipf.write(xlsx_file_path, arcname=os.path.basename(xlsx_file_path))
            if test_file_path:
                zipf.write(test_file_path, arcname=os.path.basename(test_file_path))

        return render_template('result.html',
                               id_file=os.path.basename(id_file_path),
                               xlsx_file=os.path.basename(xlsx_file_path),
                               test_file=os.path.basename(test_file_path) if test_file_path else None,
                               zip_file=os.path.basename(zip_path))

    except Exception as e:
        flash(f"Error: {str(e)}")
        return redirect(url_for('index'))

@app.route('/download/<filename>')
def download(filename):
    return send_file(os.path.join(OUTPUT_FOLDER, filename), as_attachment=True)

if __name__ == '__main__':
    app.run(debug=True)
